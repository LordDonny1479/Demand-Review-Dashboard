from __future__ import annotations

import calendar
import csv
import json
import math
import re
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
RAW_DIR = ROOT / "data" / "raw"
YOY_XLSX = RAW_DIR / "DR 07 July - YoY.xlsx"
MOM_XLSX = RAW_DIR / "DR 07 July - MoM.xlsx"
PRODUCT_XLSX = RAW_DIR / "Product List 20260629 (2).xlsx"
MARKET_XLSX = RAW_DIR / "Market List.xlsx"

OUTPUT_MODULE = ROOT / "app" / "data" / "promo-yoy-data.js"
PUBLIC_DASHBOARD_JSON = ROOT / "public" / "data" / "promo-dashboard-data.json"
OUTPUT_DASHBOARD_JSON = ROOT / "data" / "promo-yoy-dashboard.json"
OUTPUT_MOM_DASHBOARD_JSON = ROOT / "data" / "promo-mom-dashboard.json"
OUTPUT_DETAIL = ROOT / "data" / "promo-yoy-detail.csv"
OUTPUT_MOM_DETAIL = ROOT / "data" / "promo-mom-detail.csv"
OUTPUT_EXCLUDED = ROOT / "data" / "promo-yoy-excluded-rows.csv"
OUTPUT_MOM_EXCLUDED = ROOT / "data" / "promo-mom-excluded-rows.csv"
OUTPUT_DISPLAY_AUDIT = ROOT / "data" / "display-conversion-audit.csv"
OUTPUT_SUMMARY = ROOT / "data" / "dashboard-summary.json"

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
PERIOD_KEYS = ("base", "comparison")
STATUS_BY_YEAR = {
    2025: {"Closed", "Committed"},
    2026: {"Closed", "Planned", "Committed"},
}

COMPARISON_CONFIGS = {
    "yoy": {
        "label": "YoY",
        "workbook": YOY_XLSX,
        "sheets": [
            {"name": "2025", "period_key": "base", "label": "2025", "short_label": "'25", "actual_year": 2025},
            {"name": "2026", "period_key": "comparison", "label": "2026", "short_label": "'26", "actual_year": 2026},
        ],
        "period_labels": {
            "base": "2025",
            "comparison": "2026",
            "base_short": "'25",
            "comparison_short": "'26",
            "base_stat": "FY 2025 Cases",
            "comparison_stat": "FY 2026 Cases",
            "delta": "Delta",
            "delta_stat": "YoY Delta Cases",
            "pct_stat": "YoY %",
            "legend": "Grey = 2025 | Bold = 2026 | Full Year delta includes %",
        },
    },
    "mom": {
        "label": "MoM",
        "workbook": MOM_XLSX,
        "sheets": [
            {"name": "June", "period_key": "base", "label": "June", "short_label": "June", "actual_year": 2026},
            {"name": "July", "period_key": "comparison", "label": "July", "short_label": "July", "actual_year": 2026},
        ],
        "period_labels": {
            "base": "June",
            "comparison": "July",
            "base_short": "June",
            "comparison_short": "July",
            "base_stat": "June Cases",
            "comparison_stat": "July Cases",
            "delta": "Change",
            "delta_stat": "MoM Change Cases",
            "pct_stat": "MoM %",
            "legend": "Grey = June pull | Bold = July pull | Full Year change includes %",
        },
    },
}

BANNER_ORDER = []

GROUP_ORDER = [
    "Granola",
    "Hot Chocolate",
    "Sweet & Creamy",
    "Cold Beverage",
    "Syrups",
    "Instant",
    "Nespresso",
    "Roast & Ground",
    "Single Serve",
    "Soups and Hot Bowls",
    "Condensed Soup",
    "Tassimo",
    "Hot Tea",
    "Sauces",
    "Dairy",
    "Biscuit Mixes",
    "Timbits",
    "Other",
]

DATA_MODES = {
    "blended": "Blend DRPs/displays into regular-case equivalents",
    "separate": "Keep DRPs/displays separate and count each display as 1 case",
}

ROLLUP_EXCLUDED_BANNERS = {"Amazon", "Costco"}
SITE_EXCLUDED_BANNERS = {"Canada"}

VISIBLE_PRODUCT_DRILLDOWN_BANNERS = {
    "Canadian Tire",
    "Fed Coop",
    "Giant Tiger",
    "Loblaw",
    "Metro Ontario",
    "Pratts Wholesale",
    "SDM",
    "Sobeys ROC",
    "Sobeys Quebec",
    "Metro Quebec",
    "PFG",
    "Walmart",
}

DISPLAY_RE = re.compile(
    r"DISPLAY|DISPLAYER|\bDISP\b|\bDRP\b|1/2\s*DRP|HALF\s*DRP|PDQ|DISPLY",
    re.IGNORECASE,
)
PACK_RE = re.compile(
    r"(\d+(?:\.\d+)?)\s*/\s*"
    r"(\d+(?:\.\d+)?(?:\s*/\s*\d+(?:\.\d+)?)?\s*(?:GR|G|KG|ML|L|CT|EA|LB))",
    re.IGNORECASE,
)
CATEGORY_FALLBACKS = {
    "TDSSKC": {
        "pack_size": "SS KComp Unspecified",
        "planner": "Single Serve",
        "segment": "Single Serve",
    },
    "TDRG": {
        "pack_size": "R&G Unspecified",
        "planner": "Roast & Ground",
        "segment": "Roast & Ground",
    },
}


def clean(value) -> str:
    return "" if value is None else str(value).strip()


def number(value) -> float:
    if value in (None, ""):
        return 0.0
    return float(value)


def as_date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def is_display_text(*parts: str) -> bool:
    return bool(DISPLAY_RE.search(" ".join(part for part in parts if part)))


def normalize_unit(unit: str) -> str:
    return clean(unit).upper().replace(" ", "").replace("GR", "G")


def parse_pack(text: str):
    match = PACK_RE.search(clean(text).upper())
    if not match:
        return None
    return float(match.group(1)), normalize_unit(match.group(2))


def format_qty(value: float) -> str:
    return str(int(value)) if float(value).is_integer() else f"{value:g}"


def pretty_unit(unit: str) -> str:
    unit = normalize_unit(unit)
    for suffix in ("KG", "ML", "CT", "EA", "LB", "L", "G"):
        if unit.endswith(suffix):
            return f"{unit[:-len(suffix)]}{suffix.lower()}"
    return unit.lower()


def pretty_label(text: str) -> str:
    replacements = {
        "R&G": "R&G",
        "SS": "SS",
        "KCOMP": "KComp",
        "NCOMP": "NComp",
        "HB": "HB",
        "CB": "CB",
        "RTD": "RTD",
        "FVC": "FVC",
        "TDL": "TDL",
        "KCUP": "K-Cup",
    }
    parts = []
    for token in re.split(r"(\s+)", clean(text)):
        if token.isspace():
            parts.append(token)
            continue
        parts.append(replacements.get(token.upper(), token.capitalize()))
    return "".join(parts).strip()


def pretty_pack_size(pack_size: str) -> str:
    text = clean(pack_size)
    match = PACK_RE.search(text.upper())
    if not match:
        return pretty_label(text)
    prefix = text[: match.start()].strip()
    qty = format_qty(float(match.group(1)))
    unit = pretty_unit(match.group(2))
    return f"{pretty_label(prefix)} {qty}/{unit}".strip()


def load_market_map():
    if not MARKET_XLSX.exists():
        raise FileNotFoundError(f"Missing market mapping workbook: {MARKET_XLSX}")

    workbook = load_workbook(MARKET_XLSX, read_only=True, data_only=True)
    sheet = workbook.active
    market_map = {}
    banner_order = []

    for row in sheet.iter_rows(values_only=True):
        market = clean(row[0] if len(row) > 0 else "")
        banner = clean(row[1] if len(row) > 1 else "")
        if not market or not banner:
            continue
        market_map[market] = banner
        if banner not in banner_order:
            banner_order.append(banner)

    workbook.close()
    return market_map, banner_order


def active_banner_order(rows: list[dict], configured_order: list[str]) -> list[str]:
    active = {row["banner"] for row in rows}
    ordered = [banner for banner in configured_order if banner in active]
    ordered.extend(sorted(active - set(ordered)))
    return ordered


def map_banner(market: str, description: str, market_map: dict | None = None) -> str | None:
    market = clean(market)
    if market_map is not None:
        return market_map.get(market)
    return None


def infer_planner(*parts: str) -> str:
    text = " ".join(clean(part).upper() for part in parts)
    if "INSTANT" in text:
        return "Instant"
    if "NCOMP" in text or "NESPRESSO" in text:
        return "Nespresso"
    if "TASSIMO" in text:
        return "Tassimo"
    if "KCOMP" in text or "K-CUP" in text or "KCUP" in text:
        return "Single Serve"
    if "R&G" in text or "ROAST" in text or "GROUND" in text:
        return "Roast & Ground"
    if "HOT CHOC" in text or "CAPPUCCINO" in text or "CREAMER" in text:
        return "Hot Chocolate"
    if "SOUP" in text or "CHILI" in text or "HOT BOWL" in text:
        return "Soups and Hot Bowls"
    if "TEA" in text:
        return "Hot Tea"
    if "GRANOLA" in text:
        return "Granola"
    if "SYRUP" in text or "READY TO DRINK" in text or "RTD" in text or "COLD" in text:
        return "Cold Beverage"
    return "Other"


PRODUCT_GROUP_ALIASES = {
    "GRANOLA": "Granola",
    "GRANOLA BAR": "Granola",
    "HOT CHOCOLATE": "Hot Chocolate",
    "HOT CHOC & CAPPUCCINO": "Hot Chocolate",
    "SWEET & CREAMY": "Sweet & Creamy",
    "COLD BEVERAGE": "Cold Beverage",
    "ICED COFFEE & SYRUPS": "Cold Beverage",
    "SYRUPS": "Syrups",
    "INSTANT": "Instant",
    "INSTANT COFFEE": "Instant",
    "NESPRESSO": "Nespresso",
    "NESPRESSO COMPATIBLE": "Nespresso",
    "ROAST & GROUND": "Roast & Ground",
    "SINGLE SERVE": "Single Serve",
    "SINGLE SERVE (K-CUP)": "Single Serve",
    "SOUPS AND HOT BOWLS": "Soups and Hot Bowls",
    "SOUP & CHILI": "Soups and Hot Bowls",
    "CONDENSED SOUP": "Condensed Soup",
    "TASSIMO": "Tassimo",
    "HOT TEA": "Hot Tea",
    "TEA": "Hot Tea",
    "SAUCES": "Sauces",
    "DAIRY": "Dairy",
    "BISCUIT MIXES": "Biscuit Mixes",
    "TIMBITS": "Timbits",
    "OTHER": "Other",
}


def normalized_product_group(value: str) -> str:
    return PRODUCT_GROUP_ALIASES.get(clean(value).upper(), "")


def product_group_label(mpg: str, planner: str, segment: str = "") -> str:
    for value in (planner, segment):
        group = normalized_product_group(value)
        if group:
            return group

    text = f"{mpg} {planner} {segment}".upper()
    if "GRANOLA" in text:
        return "Granola"
    if "HOT CHOC" in text or "CAPPUCCINO" in text or "CREAMER" in text:
        return "Hot Chocolate"
    if "SYRUP" in text or "READY TO DRINK" in text or "RTD" in text or "COLD BEVERAGE" in text:
        return "Cold Beverage"
    if "INSTANT" in text:
        return "Instant"
    if "NCOMP" in text or "NESPRESSO" in text:
        return "Nespresso"
    if "TASSIMO" in text:
        return "Tassimo"
    if "KCOMP" in text or "K-CUP" in text or "KCUP" in text:
        return "Single Serve"
    if "R&G" in text or "ROAST" in text or "GROUND" in text:
        return "Roast & Ground"
    if "SOUP" in text or "CHILI" in text or "HOT BOWL" in text or "CONDENSED" in text:
        return "Soups and Hot Bowls"
    if "TEA" in text:
        return "Hot Tea"
    return "Other"


def display_group_label(group: str) -> str:
    return f"{group} Displays"


def family_key(rec: dict) -> str:
    planner = clean(rec.get("planner", "")).upper()
    pack_text = clean(rec.get("pack_size", "")).upper()
    item_text = clean(rec.get("item", "")).upper()
    text = " ".join(
        clean(rec.get(key, "")).upper()
        for key in ("pack_size", "item", "planner", "segment", "sub_category_1")
    )

    if "INSTANT" in planner or "INSTANT" in pack_text:
        return "instant"
    if "NESPRESSO" in planner or "NCOMP" in pack_text or "NESPRESSO" in text:
        return "nespresso"
    if "TASSIMO" in planner or "TASSIMO" in pack_text or "TASSIMO" in item_text:
        return "tassimo"
    if "HOT CHOCOLATE" in planner or "SWEET" in planner or "HOT CHOC" in pack_text:
        return "hot_chocolate"
    if "SINGLE SERVE" in planner or "KCOMP" in pack_text or "K-CUP" in text or "KCUP" in text:
        return "kcomp"
    if "ROAST" in planner or "R&G" in text or "ROAST" in text or "GROUND" in text:
        return "roast_ground"
    if "SOUP" in planner or "SOUP" in text or "CHILI" in text or "HOT BOWL" in text:
        return "soup"
    if "TEA" in planner or "TEA" in text:
        return "tea"
    if "GRANOLA" in planner or "GRANOLA" in text:
        return "granola"
    if "COLD" in planner or "SYRUP" in text or "READY TO DRINK" in text or "RTD" in text:
        return "cold_beverage"
    return clean(rec.get("planner")) or clean(rec.get("segment")) or "other"


def product_record(row) -> dict:
    pack_size_id = clean(row[6])
    pack_size = clean(row[7])
    item_id = clean(row[8])
    item = clean(row[9])
    demand_review_segment = clean(row[16]) or clean(row[5]) or clean(row[15]) or infer_planner(pack_size, item)
    rec = {
        "line_of_business": clean(row[1]),
        "brand": clean(row[3]),
        "segment": demand_review_segment,
        "source_segment": clean(row[5]),
        "pack_size_id": pack_size_id,
        "pack_size": pack_size,
        "item_id": item_id,
        "item": item,
        "sub_category_1": clean(row[15]),
        "planner": demand_review_segment,
        "lookup_source": "product_master",
    }
    rec["pack"] = parse_pack(pack_size)
    rec["is_display"] = is_display_text(pack_size_id, pack_size, item)
    return rec


def fallback_product_record(product_id: str, product: str) -> dict | None:
    pack = parse_pack(product)
    fallback = CATEGORY_FALLBACKS.get(clean(product_id).upper())
    if not pack and not fallback:
        return None
    if not pack:
        planner = fallback["planner"]
        pack_size = fallback["pack_size"]
        segment = fallback["segment"]
    else:
        planner = infer_planner(product_id, product)
        pack_size = product
        segment = planner
    rec = {
        "line_of_business": "",
        "brand": "",
        "segment": segment,
        "pack_size_id": product_id,
        "pack_size": pack_size,
        "item_id": product_id,
        "item": product,
        "sub_category_1": "",
        "planner": planner,
        "lookup_source": "demand_product_text",
        "pack": pack,
        "is_display": is_display_text(product_id, product),
        "is_unspecified_pack": bool(fallback and not pack),
    }
    return rec


def load_products():
    workbook = load_workbook(PRODUCT_XLSX, read_only=True, data_only=True)
    sheet = workbook["Products"]
    products = []
    item_by_id = {}
    pack_groups = defaultdict(list)

    for row in sheet.iter_rows(min_row=6, values_only=True):
        rec = product_record(row)
        products.append(rec)
        if rec["item_id"]:
            item_by_id[rec["item_id"]] = rec
        if rec["pack_size_id"]:
            pack_groups[rec["pack_size_id"]].append(rec)

    return products, item_by_id, pack_groups


def lookup_product(product_id: str, product: str, item_by_id, pack_groups):
    if product_id in item_by_id:
        return item_by_id[product_id]

    group = pack_groups.get(product_id)
    if group:
        display_product = is_display_text(product_id, product)
        preferred = [
            rec
            for rec in group
            if rec["is_display"] == display_product or is_display_text(product, rec["item"])
        ]
        non_display = [rec for rec in group if not rec["is_display"]]
        return (preferred or non_display or group)[0]

    return fallback_product_record(product_id, product)


def raw_demand_rows(item_by_id, pack_groups, workbook_path: Path, sheet_configs: list[dict]):
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    rows = []

    for sheet_config in sheet_configs:
        sheet_name = sheet_config["name"]
        sheet = workbook[sheet_name]
        year = sheet_config["actual_year"]
        for row_number, row in enumerate(sheet.iter_rows(min_row=7, values_only=True), start=7):
            if not any(value is not None for value in row):
                continue

            product_id = clean(row[23])
            product = clean(row[24])
            rec = lookup_product(product_id, product, item_by_id, pack_groups)
            is_display = bool(
                rec
                and is_display_text(
                    product_id,
                    product,
                    rec["pack_size_id"],
                    rec["pack_size"],
                    rec["item"],
                )
            )

            rows.append(
                {
                    "source_sheet": sheet_name,
                    "source_row": row_number,
                    "year": year,
                    "period_key": sheet_config["period_key"],
                    "period_label": sheet_config["label"],
                    "market": clean(row[0]),
                    "contract_start": as_date(row[1]),
                    "execution_start": as_date(row[2]),
                    "tls_ship_start": as_date(row[3]),
                    "contract_end": as_date(row[4]),
                    "execution_end": as_date(row[5]),
                    "description": clean(row[6]),
                    "promo_id": clean(row[7]),
                    "approval_status": clean(row[8]),
                    "promo_status": clean(row[9]),
                    "product_id": product_id,
                    "product": product,
                    "promotion_type": clean(row[28]),
                    "forecast_base_cases": number(row[34]),
                    "forecast_incremental_cases": number(row[35]),
                    "forecast_total_cases": number(row[36]),
                    "product_record": rec,
                    "is_display_product": is_display,
                }
            )

    workbook.close()
    return rows


def build_base_pack_lookup(products, demand_rows):
    master_candidates = defaultdict(Counter)
    demand_candidates = defaultdict(Counter)
    pack_size_by_key = {}

    for rec in products:
        if rec["pack"] and not rec["is_display"]:
            key = (family_key(rec), rec["pack"][1])
            master_candidates[key][rec["pack"][0]] += 1
            pack_size_by_key.setdefault((key, rec["pack"][0]), rec["pack_size"])

    for row in demand_rows:
        rec = row["product_record"]
        if not rec or not rec.get("pack") or row["is_display_product"]:
            continue
        key = (family_key(rec), rec["pack"][1])
        demand_candidates[key][rec["pack"][0]] += 1
        pack_size_by_key[(key, rec["pack"][0])] = rec["pack_size"]

    lookup = {}
    for key in set(master_candidates) | set(demand_candidates):
        source = demand_candidates.get(key) or master_candidates.get(key)
        if not source:
            continue
        pack_qty = source.most_common(1)[0][0]
        lookup[key] = {
            "pack_qty": pack_qty,
            "pack_size": pack_size_by_key.get((key, pack_qty), ""),
        }
    return lookup


def display_pack_candidates(row: dict, rec: dict):
    candidates = [
        parse_pack(row["product"]),
        parse_pack(row["product_id"]),
        parse_pack(rec.get("item", "")),
        parse_pack(rec.get("pack_size", "")),
        rec.get("pack"),
    ]
    unique = []
    seen = set()
    for candidate in candidates:
        if not candidate:
            continue
        key = (candidate[0], candidate[1])
        if key in seen:
            continue
        seen.add(key)
        unique.append(candidate)
    return unique


def equivalent_display_units(family: str, unit: str) -> list[str]:
    unit = normalize_unit(unit)
    units = [unit]
    if family == "granola" and unit == "5EA":
        units.append("5/30G")
    return units


def month_splits(start: date, end: date):
    if end < start:
        start, end = end, start

    total_days = (end - start).days + 1
    cursor = start
    splits = []
    while cursor <= end:
        last_day = calendar.monthrange(cursor.year, cursor.month)[1]
        month_end = date(cursor.year, cursor.month, last_day)
        segment_end = min(month_end, end)
        days = (segment_end - cursor).days + 1
        splits.append((cursor.month - 1, days / total_days, days, total_days))
        cursor = segment_end + timedelta(days=1)
    return splits


def round_cases(value: float) -> int:
    if value >= 0:
        return int(math.floor(value + 0.5))
    return int(math.ceil(value - 0.5))


def empty_years():
    return {period: [0.0] * 12 for period in PERIOD_KEYS}


def add_month(values, period_key: str, month_index: int, cases: float):
    values[period_key][month_index] += cases


def row_from_values(label: str, values: dict, is_group: bool = False, is_total: bool = False):
    base_months = [round_cases(value) for value in values["base"]]
    comparison_months = [round_cases(value) for value in values["comparison"]]
    base_total = round_cases(sum(values["base"]))
    comparison_total = round_cases(sum(values["comparison"]))
    return {
        "label": label,
        "is_group": is_group,
        "is_total": is_total,
        "m25": base_months,
        "m26": comparison_months,
        "fy25": base_total,
        "fy26": comparison_total,
        "base_months": base_months,
        "comparison_months": comparison_months,
        "base_total": base_total,
        "comparison_total": comparison_total,
    }


def add_values(target: dict, source: dict):
    for period in PERIOD_KEYS:
        for index, value in enumerate(source[period]):
            target[period][index] += value


def group_sort_key(group_name: str):
    display_section = 1 if group_name.endswith(" Displays") else 0
    base_name = group_name.removesuffix(" Displays")
    try:
        return (display_section, GROUP_ORDER.index(base_name), group_name)
    except ValueError:
        return (display_section, len(GROUP_ORDER), group_name)


def fy_delta(row: dict) -> int:
    return (row.get("fy26") or 0) - (row.get("fy25") or 0)


def build_product_table(rows, include_retailer_drilldown: bool = False, visible_retailer_banners: set[str] | None = None):
    group_values = defaultdict(empty_years)
    mpg_values = defaultdict(empty_years)
    retailer_values = defaultdict(empty_years)
    total_values = empty_years()

    for row in rows:
        key = (row["product_group"], row["mpg"])
        retailer_key = (row["product_group"], row["mpg"], row["banner"])
        add_month(group_values[row["product_group"]], row["period_key"], row["month_index"], row["cases"])
        add_month(mpg_values[key], row["period_key"], row["month_index"], row["cases"])
        add_month(retailer_values[retailer_key], row["period_key"], row["month_index"], row["cases"])
        add_month(total_values, row["period_key"], row["month_index"], row["cases"])

    table_rows = []
    display_section_started = False
    for group in sorted(group_values, key=group_sort_key):
        group_row = row_from_values(group, group_values[group], is_group=True)
        group_row["row_type"] = "group"
        group_row["has_children"] = True
        if group.endswith(" Displays"):
            group_row["is_display_group"] = True
            if not display_section_started:
                group_row["display_section_start"] = True
                display_section_started = True
        table_rows.append(group_row)

        children = sorted(
            (key for key in mpg_values if key[0] == group),
            key=lambda key: key[1],
        )
        for _, mpg in children:
            mpg_row = row_from_values(mpg, mpg_values[(group, mpg)])
            if include_retailer_drilldown:
                mpg_row["row_type"] = "mpg"
                mpg_row["is_mpg"] = True
                mpg_row["has_children"] = True
            table_rows.append(mpg_row)

            if include_retailer_drilldown:
                retailer_rows = []
                for _, _, banner in (key for key in retailer_values if key[0] == group and key[1] == mpg):
                    if visible_retailer_banners is not None and banner not in visible_retailer_banners:
                        continue
                    retailer_row = row_from_values(banner, retailer_values[(group, mpg, banner)])
                    retailer_row["row_type"] = "retailer"
                    retailer_row["is_retailer"] = True
                    retailer_rows.append(retailer_row)

                table_rows.extend(
                    sorted(retailer_rows, key=lambda row: (-fy_delta(row), row["label"]))
                )

    table_rows.append(row_from_values("GRAND TOTAL", total_values, is_total=True))
    return table_rows


def build_segment_table(rows, include_retailer_drilldown: bool = False, visible_retailer_banners: set[str] | None = None):
    group_values = defaultdict(empty_years)
    retailer_values = defaultdict(empty_years)
    total_values = empty_years()

    for row in rows:
        retailer_key = (row["product_group"], row["banner"])
        add_month(group_values[row["product_group"]], row["period_key"], row["month_index"], row["cases"])
        add_month(retailer_values[retailer_key], row["period_key"], row["month_index"], row["cases"])
        add_month(total_values, row["period_key"], row["month_index"], row["cases"])

    table_rows = []
    display_section_started = False
    for group in sorted(group_values, key=group_sort_key):
        retailer_rows = []
        if include_retailer_drilldown:
            for _, banner in (key for key in retailer_values if key[0] == group):
                if visible_retailer_banners is not None and banner not in visible_retailer_banners:
                    continue
                retailer_row = row_from_values(banner, retailer_values[(group, banner)])
                retailer_row["row_type"] = "retailer"
                retailer_row["is_retailer"] = True
                retailer_row["parent_level"] = "group"
                retailer_rows.append(retailer_row)

        group_row = row_from_values(group, group_values[group], is_group=True)
        group_row["row_type"] = "group"
        group_row["has_children"] = bool(retailer_rows)
        if group.endswith(" Displays"):
            group_row["is_display_group"] = True
            if not display_section_started:
                group_row["display_section_start"] = True
                display_section_started = True
        table_rows.append(group_row)

        table_rows.extend(sorted(retailer_rows, key=lambda row: (-fy_delta(row), row["label"])))

    table_rows.append(row_from_values("GRAND TOTAL", total_values, is_total=True))
    return table_rows


def build_retailer_rollup(rows):
    retailer_values = {banner: empty_years() for banner in BANNER_ORDER}
    total_values = empty_years()

    for row in rows:
        add_month(retailer_values[row["banner"]], row["period_key"], row["month_index"], row["cases"])
        add_month(total_values, row["period_key"], row["month_index"], row["cases"])

    table_rows = [
        row_from_values(banner, retailer_values[banner], is_group=True)
        for banner in BANNER_ORDER
        if sum(retailer_values[banner]["base"]) or sum(retailer_values[banner]["comparison"])
    ]
    table_rows.append(row_from_values("GRAND TOTAL", total_values, is_total=True))
    return table_rows, total_values


def conversion_for_row(row: dict, base_lookup: dict):
    rec = row["product_record"]
    unit_type = "CASE"
    conversion = 1.0
    base_pack_size = rec["pack_size"]
    conversion_note = "Reported as regular cases"
    display_pack = None
    base = None

    if row["is_display_product"]:
        unit_type = "DRP"
        candidates = display_pack_candidates(row, rec)
        for candidate in candidates:
            family = family_key(rec)
            for unit in equivalent_display_units(family, candidate[1]):
                key = (family, unit)
                if key in base_lookup:
                    display_pack = candidate
                    base = base_lookup[key]
                    break
            if base:
                break
        if display_pack is None and candidates:
            display_pack = candidates[0]
            base = base_lookup.get((family_key(rec), display_pack[1]))

        if display_pack and base:
            conversion = display_pack[0] / base["pack_qty"]
            base_pack_size = base["pack_size"] or rec["pack_size"]
            conversion_note = (
                f"Display pack {format_qty(display_pack[0])}/{display_pack[1]} "
                f"converted to regular case pack {format_qty(base['pack_qty'])}/{display_pack[1]}"
            )
        else:
            conversion_note = "Display product without matching regular pack; left as reported cases"

    return {
        "unit_type": unit_type,
        "conversion": conversion,
        "base_pack_size": base_pack_size,
        "conversion_note": conversion_note,
        "display_pack_qty": "" if display_pack is None else display_pack[0],
        "display_pack_unit": "" if display_pack is None else display_pack[1],
    }


def build_dashboard_from_rows(rows: list[dict]):
    rollup_rows = [
        row
        for row in rows
        if row["banner"] not in ROLLUP_EXCLUDED_BANNERS
    ]
    rollup_ret, total_values = build_retailer_rollup(rollup_rows)
    all_retailer_totals, _ = build_retailer_rollup(rows)
    stats = build_stats(total_values)
    stats["banners"] = len({row["banner"] for row in rollup_rows})
    return {
        "rollup_ret": rollup_ret,
        "rollup_grp": build_product_table(
            rollup_rows,
            include_retailer_drilldown=True,
            visible_retailer_banners=VISIBLE_PRODUCT_DRILLDOWN_BANNERS,
        ),
        "rollup_segment": build_segment_table(
            rollup_rows,
            include_retailer_drilldown=True,
            visible_retailer_banners=VISIBLE_PRODUCT_DRILLDOWN_BANNERS,
        ),
        "retailer_totals": [
            row
            for row in all_retailer_totals
            if not row.get("is_total")
        ],
        "retailers": {
            banner: build_product_table([row for row in rows if row["banner"] == banner])
            for banner in BANNER_ORDER
        },
        "stats": stats,
    }


def transform_comparison(comparison_key: str, config: dict, demand_rows: list[dict], base_lookup: dict, market_map: dict):
    mode_rows = {mode: [] for mode in DATA_MODES}
    detail_rows = []
    included_split_keys = set()
    included_source_rows = set()
    excluded_rows = []
    display_audit = {}

    for row in demand_rows:
        year = row["year"]
        banner = map_banner(row["market"], row["description"], market_map)
        rec = row["product_record"]
        reason = None

        if row["promo_status"] not in STATUS_BY_YEAR[year]:
            reason = "Promo status outside dashboard methodology"
        elif row["forecast_incremental_cases"] <= 0:
            reason = "Fcst Inc Cases is not positive"
        elif not banner:
            reason = "Market is outside the supplied market list"
        elif banner in SITE_EXCLUDED_BANNERS:
            reason = "Retailer is excluded from this dashboard"
        elif not rec or (not rec.get("pack") and not rec.get("is_unspecified_pack")):
            reason = "Product cannot be mapped to an MPG pack size"
        elif not row["execution_start"] or not row["execution_end"]:
            reason = "Missing execution start or end date"

        if reason:
            excluded_rows.append(excluded_row(row, banner, reason, comparison_key))
            continue

        conversion = conversion_for_row(row, base_lookup)
        blended_mpg = pretty_pack_size(conversion["base_pack_size"])
        blended_product_group = product_group_label(blended_mpg, rec["planner"], rec["segment"])
        separate_mpg = pretty_pack_size(rec["pack_size"])
        separate_product_group = product_group_label(separate_mpg, rec["planner"], rec["segment"])
        source_cases = row["forecast_incremental_cases"]
        converted_cases = source_cases * conversion["conversion"]

        if conversion["unit_type"] == "DRP":
            separate_product_group = display_group_label(blended_product_group)

        if conversion["unit_type"] == "DRP":
            converted = "without matching regular pack" not in conversion["conversion_note"]
            audit_key = (
                comparison_key,
                row["product_id"],
                row["product"],
                blended_mpg,
                separate_mpg,
                conversion["conversion"],
                conversion["conversion_note"],
            )
            display_audit[audit_key] = {
                "comparison_key": comparison_key,
                "product_id": row["product_id"],
                "product": row["product"],
                "blended_mpg": blended_mpg,
                "separate_mpg": separate_mpg,
                "blended_product_group": blended_product_group,
                "separate_product_group": separate_product_group,
                "cases_per_display": round(conversion["conversion"], 6),
                "converted": "yes" if converted else "no",
                "conversion_note": conversion["conversion_note"],
            }

        for month_index, weight, split_days, total_days in month_splits(row["execution_start"], row["execution_end"]):
            included_split_keys.add((row["source_sheet"], row["source_row"], month_index))
            included_source_rows.add((row["source_sheet"], row["source_row"]))
            mode_definitions = [
                ("blended", blended_product_group, blended_mpg, converted_cases),
                ("separate", separate_product_group, separate_mpg, source_cases),
            ]
            for data_mode, product_group, mpg, mode_cases in mode_definitions:
                month_cases = mode_cases * weight
                mode_rows[data_mode].append(
                    {
                        "banner": banner,
                        "market": row["market"],
                        "year": year,
                        "period_key": row["period_key"],
                        "period_label": row["period_label"],
                        "month": MONTHS[month_index],
                        "month_index": month_index,
                        "product_group": product_group,
                        "mpg": mpg,
                        "cases": month_cases,
                    }
                )
                detail_rows.append({
                    "comparison_key": comparison_key,
                    "data_mode": data_mode,
                    "banner": banner,
                    "market": row["market"],
                    "year": year,
                    "period": row["period_label"],
                    "month": MONTHS[month_index],
                    "product_group": product_group,
                    "mpg": mpg,
                    "product_id": row["product_id"],
                    "product": row["product"],
                    "promo_id": row["promo_id"],
                    "promo_status": row["promo_status"],
                    "promotion_type": row["promotion_type"],
                    "execution_start": row["execution_start"].isoformat(),
                    "execution_end": row["execution_end"].isoformat(),
                    "source_fcst_inc_cases": round(source_cases, 6),
                    "unit_type": conversion["unit_type"],
                    "cases_per_display": "" if conversion["unit_type"] == "CASE" else round(conversion["conversion"], 6),
                    "converted_fcst_inc_cases": round(converted_cases, 6),
                    "mode_fcst_inc_cases": round(mode_cases, 6),
                    "prorate_weight": round(weight, 8),
                    "execution_days_in_month": split_days,
                    "execution_days_total": total_days,
                    "month_cases": round(month_cases, 6),
                    "source_sheet": row["source_sheet"],
                    "source_row": row["source_row"],
                    "conversion_note": conversion["conversion_note"],
                })

    return {
        "mode_rows": mode_rows,
        "detail_rows": detail_rows,
        "excluded_rows": excluded_rows,
        "display_audit": display_audit,
        "included_split_keys": included_split_keys,
        "included_source_rows": included_source_rows,
        "source_rows_read": len(demand_rows),
        "workbook": config["workbook"].name,
    }


def comparison_summary(config: dict, transformed: dict, dashboard_modes: dict):
    display_audit = transformed["display_audit"]
    return {
        "source_workbook": config["workbook"].name,
        "periods": [
            {
                "sheet": sheet["name"],
                "label": sheet["label"],
                "actual_year": sheet["actual_year"],
                "status_filter": sorted(STATUS_BY_YEAR[sheet["actual_year"]]),
            }
            for sheet in config["sheets"]
        ],
        "period_labels": config["period_labels"],
        "source_rows_read": transformed["source_rows_read"],
        "included_source_month_rows": len(transformed["included_split_keys"]),
        "included_source_rows": len(transformed["included_source_rows"]),
        "excluded_rows": len(transformed["excluded_rows"]),
        "display_products_reviewed": len(display_audit),
        "display_products_converted": sum(1 for row in display_audit.values() if row["converted"] == "yes"),
        "unconverted_display_products": sum(1 for row in display_audit.values() if row["converted"] == "no"),
        "detail_rows": len(transformed["detail_rows"]),
        "mode_totals": {
            mode: data["stats"]
            for mode, data in dashboard_modes.items()
        },
    }


def build_outputs():
    global BANNER_ORDER

    market_map, configured_banner_order = load_market_map()
    products, item_by_id, pack_groups = load_products()
    demand_rows_by_comparison = {
        key: raw_demand_rows(item_by_id, pack_groups, config["workbook"], config["sheets"])
        for key, config in COMPARISON_CONFIGS.items()
    }
    all_demand_rows = [
        row
        for demand_rows in demand_rows_by_comparison.values()
        for row in demand_rows
    ]
    base_lookup = build_base_pack_lookup(products, all_demand_rows)

    transformed = {
        key: transform_comparison(key, config, demand_rows_by_comparison[key], base_lookup, market_map)
        for key, config in COMPARISON_CONFIGS.items()
    }

    BANNER_ORDER = active_banner_order(
        [
            row
            for result in transformed.values()
            for rows in result["mode_rows"].values()
            for row in rows
        ],
        configured_banner_order,
    )

    comparison_dashboards = {}
    comparison_summaries = {}
    for key, config in COMPARISON_CONFIGS.items():
        dashboard_modes = {
            mode: build_dashboard_from_rows(rows)
            for mode, rows in transformed[key]["mode_rows"].items()
        }
        comparison_dashboards[key] = {
            "label": config["label"],
            "period_labels": config["period_labels"],
            "mode_labels": DATA_MODES,
            "modes": dashboard_modes,
            **dashboard_modes["blended"],
        }
        comparison_summaries[key] = comparison_summary(config, transformed[key], dashboard_modes)

    yoy_dashboard = comparison_dashboards["yoy"]
    yoy_summary = comparison_summaries["yoy"]
    dashboard = {
        "banner_order": BANNER_ORDER,
        "default_mode": "blended",
        "mode_labels": DATA_MODES,
        "comparisons": comparison_dashboards,
        "modes": yoy_dashboard["modes"],
        **yoy_dashboard["modes"]["blended"],
    }

    summary = {
        "generated_from": {
            "demand_workbook": YOY_XLSX.name,
            "yoy_workbook": YOY_XLSX.name,
            "mom_workbook": MOM_XLSX.name,
            "product_workbook": PRODUCT_XLSX.name,
            "market_workbook": MARKET_XLSX.name,
        },
        "methodology": {
            "volume_source": "Product-level Fcst Inc Cases",
            "row_filter": "Fcst Inc Cases > 0",
            "date_method": "Execution Start through Execution End, prorated by inclusive execution days per calendar month",
            "year_status_filter": {
                "2025": sorted(STATUS_BY_YEAR[2025]),
                "2026": sorted(STATUS_BY_YEAR[2026]),
            },
            "display_method": "Toggleable: blended mode converts display/DRP/PDQ pack sizes to equivalent regular cases; separate mode keeps display/DRP rows separate and counts each as 1 case",
            "product_level": "Product group is sourced from Product List column Q (Demand Review Planner); MPG pack-size level combines individual flavours",
            "banner_scope": BANNER_ORDER,
            "market_mapping": "Retailer/customer names are mapped from Market List.xlsx",
            "mom_comparison": "MoM compares the July pull against the June pull from DR 07 July - MoM.xlsx using the same product, market, status, date, and display-conversion methodology",
            "site_excluded_banners": sorted(SITE_EXCLUDED_BANNERS),
            "rollup_excluded_banners": sorted(ROLLUP_EXCLUDED_BANNERS),
            "product_drilldown_visible_banners": sorted(VISIBLE_PRODUCT_DRILLDOWN_BANNERS),
        },
        "comparisons": comparison_summaries,
        "source_rows_read": yoy_summary["source_rows_read"],
        "data_modes": DATA_MODES,
        "included_source_month_rows": yoy_summary["included_source_month_rows"],
        "included_source_rows": yoy_summary["included_source_rows"],
        "excluded_rows": yoy_summary["excluded_rows"],
        "display_products_reviewed": yoy_summary["display_products_reviewed"],
        "display_products_converted": yoy_summary["display_products_converted"],
        "unconverted_display_products": yoy_summary["unconverted_display_products"],
        "detail_rows": yoy_summary["detail_rows"],
        "mode_totals": yoy_summary["mode_totals"],
        "total_cases_2025": yoy_dashboard["modes"]["blended"]["stats"]["fy25"],
        "total_cases_2026": yoy_dashboard["modes"]["blended"]["stats"]["fy26"],
        "delta_cases": yoy_dashboard["modes"]["blended"]["stats"]["delta"],
        "delta_pct": yoy_dashboard["modes"]["blended"]["stats"]["delta_pct"],
    }

    OUTPUT_MODULE.parent.mkdir(parents=True, exist_ok=True)
    PUBLIC_DASHBOARD_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_DASHBOARD_JSON.parent.mkdir(parents=True, exist_ok=True)
    PUBLIC_DASHBOARD_JSON.write_text(
        json.dumps({"MONTHS": MONTHS, "RAW": dashboard, "META": summary}, separators=(",", ":")),
        encoding="utf-8",
    )
    OUTPUT_DASHBOARD_JSON.write_text(json.dumps(dashboard, indent=2), encoding="utf-8")
    OUTPUT_MOM_DASHBOARD_JSON.write_text(json.dumps(comparison_dashboards["mom"], indent=2), encoding="utf-8")
    OUTPUT_SUMMARY.write_text(json.dumps(summary, indent=2), encoding="utf-8")
    OUTPUT_MODULE.write_text(module_text(dashboard, summary), encoding="utf-8")

    write_csv(OUTPUT_DETAIL, detail_fieldnames(), transformed["yoy"]["detail_rows"])
    write_csv(OUTPUT_MOM_DETAIL, detail_fieldnames(), transformed["mom"]["detail_rows"])
    write_csv(OUTPUT_EXCLUDED, excluded_fieldnames(), transformed["yoy"]["excluded_rows"])
    write_csv(OUTPUT_MOM_EXCLUDED, excluded_fieldnames(), transformed["mom"]["excluded_rows"])
    all_display_audit = [
        row
        for result in transformed.values()
        for row in result["display_audit"].values()
    ]
    write_csv(
        OUTPUT_DISPLAY_AUDIT,
        [
            "comparison_key",
            "product_id",
            "product",
            "blended_product_group",
            "blended_mpg",
            "separate_product_group",
            "separate_mpg",
            "cases_per_display",
            "converted",
            "conversion_note",
        ],
        sorted(all_display_audit, key=lambda row: (row["comparison_key"], row["blended_product_group"], row["blended_mpg"], row["product_id"])),
    )

    return summary


def build_stats(total_values: dict):
    base_total = round_cases(sum(total_values["base"]))
    comparison_total = round_cases(sum(total_values["comparison"]))
    delta = comparison_total - base_total
    delta_pct = round((delta / abs(base_total)) * 100, 1) if base_total else None
    return {
        "fy25": base_total,
        "fy26": comparison_total,
        "base_total": base_total,
        "comparison_total": comparison_total,
        "delta": delta,
        "delta_pct": delta_pct,
        "banners": len(BANNER_ORDER),
    }


def module_text(dashboard: dict, summary: dict) -> str:
    return (
        "// Generated by scripts/build_dashboard_data.py. Do not edit by hand.\n"
        "export const DASHBOARD_DATA_URL = \"/data/promo-dashboard-data.json\";\n"
    )


def excluded_row(row: dict, banner: str | None, reason: str, comparison_key: str):
    return {
        "comparison_key": comparison_key,
        "reason": reason,
        "banner": banner or "",
        "market": row["market"],
        "year": row["year"],
        "period": row["period_label"],
        "product_id": row["product_id"],
        "product": row["product"],
        "promo_id": row["promo_id"],
        "promo_status": row["promo_status"],
        "forecast_incremental_cases": row["forecast_incremental_cases"],
        "execution_start": "" if not row["execution_start"] else row["execution_start"].isoformat(),
        "execution_end": "" if not row["execution_end"] else row["execution_end"].isoformat(),
        "source_sheet": row["source_sheet"],
        "source_row": row["source_row"],
        "description": row["description"],
    }


def detail_fieldnames():
    return [
        "comparison_key",
        "data_mode",
        "banner",
        "market",
        "year",
        "period",
        "month",
        "product_group",
        "mpg",
        "product_id",
        "product",
        "promo_id",
        "promo_status",
        "promotion_type",
        "execution_start",
        "execution_end",
        "source_fcst_inc_cases",
        "unit_type",
        "cases_per_display",
        "converted_fcst_inc_cases",
        "mode_fcst_inc_cases",
        "prorate_weight",
        "execution_days_in_month",
        "execution_days_total",
        "month_cases",
        "source_sheet",
        "source_row",
        "conversion_note",
    ]


def excluded_fieldnames():
    return [
        "comparison_key",
        "reason",
        "banner",
        "market",
        "year",
        "period",
        "product_id",
        "product",
        "promo_id",
        "promo_status",
        "forecast_incremental_cases",
        "execution_start",
        "execution_end",
        "source_sheet",
        "source_row",
        "description",
    ]


def write_csv(path: Path, fieldnames: list[str], rows: list[dict]):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


if __name__ == "__main__":
    print(json.dumps(build_outputs(), indent=2))
